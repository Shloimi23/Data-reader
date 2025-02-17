import time
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

RECIPIENT_EMAIL = "sh5684770@gmail.com"
PREVIOUS_FILE = "previous_data.xlsx"


#Function that reads data from the site and returns an html object from the site and sorts data from it
def reading_from_the_link(url):
    response = requests.get(url)  # Reading data from the site
    soup = BeautifulSoup(response.text, "html.parser")  # HTML processing
    rows = soup.find_all("tr", {"class": "data"})  # Find all desired rows

    data = []

    for row in rows:
        name = row.find("td", {"class": "qName"}).find("a").text.strip()
        price = row.find("td", {"class": "last"}).text.strip()
        time_element = row.find("td", {"class": "lrtime"})
        time = time_element.text.strip() if time_element else "לא זמין"

        # Retrieving the daily change
        change_element = row.find("span", {"class": "red bgr"}) or \
                         row.find("span", {"class": "green bgg"}) or \
                         row.find("span", {"class": "nopchange"}) or \
                         row.find("span", {"class": "green bgg bggreen"}) or \
                         row.find("span", {"class": "red bgr bgred"})
        #print(f"שורת שינוי יומי: {row.find('td', {'class': 'changeP'})}")

        change_value = "0%"
        if change_element and change_element.text.strip():
            change_value = change_element.text.strip()

        #print(f"שם מניה: {name}, מחיר: {price}, סוג מסחר: {time}, שינוי יומי: {change_value}")
        data.append({"שם מניה": name, "מחיר": price, "שעה": time, "שינוי יומי": change_value})

    create_excel_file(data)

#Function that creates an Excel file
def create_excel_file(data):
    df = pd.DataFrame(data)
    current_time = datetime.now().strftime("%H_%M")
    file_name = f"snp_{current_time}.xlsx"
    df.to_excel(file_name, index=False, engine="openpyxl")

    sort_excel_by_change(file_name)

    if os.path.exists(PREVIOUS_FILE):
        previous_df = pd.read_excel(PREVIOUS_FILE)
        compare_and_color_changes(file_name, df, previous_df)

    # Save the current file as the previous file
    df.to_excel(PREVIOUS_FILE, index=False, engine="openpyxl")
    print(f"הנתונים נשמרו בהצלחה בקובץ {file_name}")
    print(f"קובץ האקסל נשמר בתיקייה: {os.getcwd()}")

# Function that colors the rows according to filtering
def compare_and_color_changes(file_name, current_df, previous_df):
    wb = load_workbook(file_name)
    ws = wb.active

    # Go over all lines in the current file
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
                                  start=2):
        current_name = row[0].value
        current_change = row[3].value

        if current_name in previous_df["שם מניה"].values:
            previous_change = previous_df.loc[previous_df["שם מניה"] == current_name, "שינוי יומי"].values[0]

            current_change = float(current_change.replace("%", "").strip()) if current_change else 0.0
            previous_change = float(previous_change.replace("%", "").strip()) if previous_change else 0.0

            try:
                if abs(current_change - previous_change) > 1 :
                    fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    for cell in row:
                        cell.fill = fill
                        #print(f"השורה {row_idx} נצבעה (שם מניה: {current_name})")
            except ValueError:
                continue

    wb.save(file_name)
    send_email_with_attachment(file_name)
    print(f"Workbook saved successfully: {file_name}")



#Function that sorts the rows by the daily change value
def sort_excel_by_change(file_name):
    df = pd.read_excel(file_name)
    df["שינוי יומי"] = df["שינוי יומי"].str.replace("%", "").astype(float)
    df = df.sort_values(by="שינוי יומי", ascending=False)
    df["שינוי יומי"] = df["שינוי יומי"].astype(str) + "%"

    # Save the sorted table in the original file
    df.to_excel(file_name, index=False, engine="openpyxl")
    print(f"הקובץ מוין לפי עמודת 'שינוי יומי' ונשמר בקובץ המקורי: {file_name}")

#Function to send an email with the file
def send_email_with_attachment(file_path):
    # Email sending settings
    sender_email = "tzviprus@gmail.com"  # Sender address
    sender_password = "uzil nvms yxmk ltoi"  # Sender's app password
    subject = "קובץ מניות חדש נוצר"
    body = "שלום,\n\nמצורף קובץ המניות שנוצר כעת.\n\nבברכה,\nהמערכת"

    # Create an email message
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = RECIPIENT_EMAIL
    msg["Subject"] = subject

    # Add the message body
    msg.attach(MIMEText(body, "plain"))

    # Add the attachment
    with open(file_path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(file_path)}", # Name of the attached file
        )
        msg.attach(part)

    # Sending the email
    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()  #Secure connection
            server.login(sender_email, sender_password)
            server.send_message(msg)
            print(f"המייל נשלח בהצלחה לכתובת {RECIPIENT_EMAIL}")
    except Exception as e:
        print(f"שגיאה בשליחת המייל: {e}")


#Function that causes the program to run every hour
def waiting_for_a_full_hour():
    now = datetime.now()
    next_hour = (now + timedelta(hours=1)).replace(minute=0,second=0,microsecond=0)
    wait_time = (next_hour - now).total_seconds()
    print(f"ממתין {wait_time / 60:.2f} דקות עד לשעה {next_hour.strftime('%H:%M')}")
    time.sleep(wait_time)


# The main function that causes the program to start running
def main(url):
    while True:
        print(f"מתחיל את התהליך בשעה: {datetime.now().strftime('%H:%M:%S')}")
        waiting_for_a_full_hour()
        reading_from_the_link(url)
        #time.sleep(600)

# Page URL
url = "https://www.globes.co.il/portal/instrument.aspx?instrumentid=373853&feeder=1&mode=composition&showAll=true#jt40991"
main(url)

